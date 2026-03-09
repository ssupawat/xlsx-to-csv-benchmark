"""
01_generate_data.py
Generate test XLSX files for benchmarking.

Files created:
  data/small.xlsx   -   1,000 rows x 20 cols (~0.2 MB)
  data/medium.xlsx  -  50,000 rows x 20 cols (~9.5 MB)
  data/large.xlsx   - 200,000 rows x 20 cols (~38  MB)
  data/multisheet.xlsx - 5 sheets x 200,000 rows x 10 cols (~57 MB)

NOTE: For the "1M rows x 5 sheets" scenario described in the analysis,
      each sheet takes ~70s to generate with xlsxwriter.
      Use 02_generate_large.py for that (runs per-sheet, saves separately).
"""

import os
import random
import string
import time
import openpyxl
import xlsxwriter

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")
os.makedirs(DATA_DIR, exist_ok=True)


# ── helpers ───────────────────────────────────────────────────────────────────

def random_row(cols=20):
    row = []
    for j in range(cols):
        match j % 4:
            case 0: row.append(random.randint(0, 1_000_000))
            case 1: row.append(round(random.uniform(0, 10_000), 4))
            case 2: row.append("".join(random.choices(string.ascii_letters, k=random.randint(5, 30))))
            case 3: row.append(random.choice([True, False, None]))
    return row


def generate_openpyxl(path, rows, cols=20):
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet()
    ws.append([f"col_{i}" for i in range(cols)])
    for _ in range(rows):
        ws.append(random_row(cols))
    wb.save(path)
    mb = os.path.getsize(path) / 1024 / 1024
    print(f"  {os.path.basename(path)}: {rows:,} rows x {cols} cols => {mb:.2f} MB")


def generate_multisheet_xlsxwriter(path, sheets=5, rows_per_sheet=200_000, cols=10):
    wb = xlsxwriter.Workbook(path, {"constant_memory": True})
    for s in range(sheets):
        ws = wb.add_worksheet(f"Sheet{s + 1}")
        for i in range(cols):
            ws.write(0, i, f"col_{i}")
        for r in range(1, rows_per_sheet + 1):
            ws.write_row(r, 0, [
                random.randint(0, 999_999),
                round(random.uniform(0, 9_999), 2),
                "".join(random.choices(string.ascii_letters, k=12)),
                r % 2,
                random.randint(0, 100),
                round(random.uniform(0, 500), 3),
                str(random.randint(10_000_000, 99_999_999)),
                random.randint(0, 9_999),
                round(random.uniform(0, 1), 5),
                r % 3,
            ])
        print(f"    Sheet{s + 1} written")
    wb.close()
    mb = os.path.getsize(path) / 1024 / 1024
    print(f"  {os.path.basename(path)}: {sheets} sheets x {rows_per_sheet:,} rows => {mb:.2f} MB")


# ── main ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    t0 = time.time()
    print("Generating benchmark data...\n")

    print("Single-sheet files (openpyxl):")
    generate_openpyxl(os.path.join(DATA_DIR, "small.xlsx"),   1_000)
    generate_openpyxl(os.path.join(DATA_DIR, "medium.xlsx"),  50_000)
    generate_openpyxl(os.path.join(DATA_DIR, "large.xlsx"),  200_000)

    print("\nMulti-sheet file (xlsxwriter, constant_memory):")
    generate_multisheet_xlsxwriter(os.path.join(DATA_DIR, "multisheet.xlsx"))

    print(f"\nDone in {time.time() - t0:.1f}s")
