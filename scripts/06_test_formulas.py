"""
06_test_formulas.py
Test how each tool handles Excel formulas.

Key finding: ALL tools read cached values only — none evaluate formulas.
  - Files saved by Excel / xlsxwriter WITH cached values → ✅ correct output
  - Files saved by openpyxl WITHOUT cached values        → ❌ empty cells

Run this to verify formula behaviour on your own xlsx files.
"""

import os
import tempfile
import xlsxwriter

DATA_DIR = os.path.join(os.path.dirname(__file__), "..", "data")
os.makedirs(DATA_DIR, exist_ok=True)


def create_formula_xlsx(path: str):
    """Create a workbook with formulas AND their cached values."""
    wb = xlsxwriter.Workbook(path)
    ws = wb.add_worksheet()
    ws.write_row(0, 0, ["price", "qty", "total", "tax", "grand_total"])
    data = [(100, 5), (250, 3)]
    for r, (price, qty) in enumerate(data, 1):
        total = price * qty
        tax = round(total * 0.07, 2)
        grand = round(total + tax, 2)
        ws.write(r, 0, price)
        ws.write(r, 1, qty)
        ws.write_formula(r, 2, f"=A{r+1}*B{r+1}", None, total)
        ws.write_formula(r, 3, f"=C{r+1}*0.07",   None, tax)
        ws.write_formula(r, 4, f"=C{r+1}+D{r+1}", None, grand)
    wb.close()


def read_calamine(path: str) -> list[list]:
    from python_calamine import CalamineWorkbook
    wb = CalamineWorkbook.from_path(path)
    return list(wb.get_sheet_by_index(0).to_python())


def read_xlsx2csv(path: str) -> str:
    from xlsx2csv import Xlsx2csv
    with tempfile.NamedTemporaryFile(suffix=".csv", delete=False, mode="w") as tf:
        out = tf.name
    Xlsx2csv(path, outputencoding="utf-8").convert(out)
    with open(out) as f:
        content = f.read()
    os.unlink(out)
    return content


def read_openpyxl(path: str) -> list[tuple]:
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    return list(wb.active.iter_rows(values_only=True))


if __name__ == "__main__":
    path = os.path.join(DATA_DIR, "formulas_cached.xlsx")
    create_formula_xlsx(path)
    print(f"Created {path}\n")

    print("=" * 55)
    print("  Formula Handling Test")
    print("  (file has formulas WITH cached values via xlsxwriter)")
    print("=" * 55)

    print("\n── python-calamine ──────────────────────────")
    for row in read_calamine(path):
        print(" ", row)

    print("\n── xlsx2csv ─────────────────────────────────")
    print(read_xlsx2csv(path))

    print("── openpyxl (data_only=True) ────────────────")
    for row in read_openpyxl(path):
        print(" ", row)

    print("\n⚠️  All tools read CACHED values, not live formula results.")
    print("   If the file was saved by openpyxl without setting cache,")
    print("   all formula cells will be empty / None.")
    print("   For live evaluation, use LibreOffice or the `formulas` library.")
